from openpyxl import load_workbook
from conf import project_path
class DoExcel():
    def __init__(self,file_name,sheet):
        self.file_name=file_name
        self.sheet=sheet
    def no_reg_tel(self):
        wb = load_workbook(self.file_name)
        sheet = wb['init']
        no_reg_tel=sheet.cell(1,2).value
        return no_reg_tel
    def read_excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        test_data=[]
        no_reg_tel=self.no_reg_tel()
        for i in range(2,sheet.max_row+1):
            list_data=[]
            for j in range(1,8):
                if j==6:
                    param = eval(sheet.cell(i, 6).value)
                    print(type(param))
                    if param['mobile']=='first_tel':
                        param['mobile']=no_reg_tel
                    list_data.append(param)

                else:
                        list_data.append(sheet.cell(i,j).value)
                list_data.append(sheet.cell(i, j).value)
            test_data.append(list_data)
            # self.updata_tel(str(int(no_reg_tel)+1))
        return test_data
    def write_excel(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        sheet.cell(row, 8).value = actual
        sheet.cell(row, 9).value = result
        wb.save(self.file_name)

if __name__ == '__main__':
    #实例化
    path=project_path.test_data_path
    t=DoExcel(path, "Sheet1")
    print(t.read_excel())












