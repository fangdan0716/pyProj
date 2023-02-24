from openpyxl import load_workbook
from conf import project_path
class DoExcel():
    def __init__(self,file_name,sheet):
        self.file_name=file_name
        self.sheet=sheet
    def read_excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        id_data=[]
        for i in range(2, sheet.max_row + 1):
            id_data.append(sheet.cell(i, 1).value)
        return id_data
    def read_excel1(self):
        wb=load_workbook(self.file_name)
        ws = wb.active
        sheet=wb[self.sheet]
        id_data=[]
        for i in range(2, sheet.max_row + 1):
            for j in  range(1,ws.max_column + 1):
                id_data.append(sheet.cell(i, j).value)
        return id_data
    def write_excel(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        sheet.cell(row, 8).value = actual
        sheet.cell(row, 9).value = result
        wb.save(self.file_name)

if __name__ == '__main__':
    #实例化
    path=project_path.id_data_path
    print(path)
    t=DoExcel(path, "Sheet1")
    print(t.read_excel1())
    list1 =t.read_excel1()
    print(type(list1[0]))












