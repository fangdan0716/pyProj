import uuid

from openpyxl import load_workbook

from comm.get_uuid import get_uuid
from conf import project_path
import time
class DoExcel():
    def __init__(self,file_name,sheet):
        self.file_name=file_name
        self.sheet=sheet
    def read_itemsAndScore(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        Items_data=[]
        scores_data=[]
        for i in range(2, sheet.max_row + 1):
            Items_data.append(sheet.cell(i, 1).value)
            scores_data.append(sheet.cell(i, 3).value)
        return Items_data,scores_data
    def read_testers(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        testers=[]
        for i in range(2, sheet.max_row + 1):
            testers.append(sheet.cell(i, 1).value)
        return testers
    def read_types(self):
        wb=load_workbook(self.file_name)
        ws = wb.active
        sheet=wb[self.sheet]
        types=[]
        for i in range(2, sheet.max_row + 1):
            for j in  range(1,3):
                types.append(sheet.cell(i, j).value)
        return types
    def write_excel(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        sheet.cell(row, 8).value = actual
        sheet.cell(row, 9).value = result
        wb.save(self.file_name)
    def writeParmtoExcel(self,row):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        testers_date = DoExcel(project_path.do_excel_new3_path, "Sheet1").read_testers()
        items_scores = DoExcel(project_path.do_excel_new3_path, "Sheet2").read_itemsAndScore()
        typedata = DoExcel(project_path.do_excel_new3_path, "Sheet3").read_types()
        parm = {"testType": "test_type","score": "scores_106","itemId": "item_id","testUnicode": "","testerId": "tester_id","keyNum": "1","evaType": "eva_Type","uuid": "uuid","faceUnicode": "","startDate": "startDate","endDate": "endDate","deviceUnicode": ""}
        for k in range(0,len(testers_date)):
            for j in range(0, len(items_scores[0])):
                uuid = str(get_uuid())
                parm["testType"] = str(typedata[0])
                parm["score"] = items_scores[1][j]
                parm["itemId"] = items_scores[0][j]
                parm["testUnicode"] = ""
                parm["testerId"] = testers_date[k]
                parm["keyNum"] = "1"
                parm["evaType"] = str(typedata[1])
                parm["uuid"] = uuid
                parm["faceUnicode"] = ""
                parm["startDate"] = time.strftime("%Y-%m-%d %H:%M:%S")
                parm["endDate"] = time.strftime("%Y-%m-%d %H:%M:%S")
                parm["deviceUnicode"] = ""
                sheet.cell(row,1).value = row - 1
                sheet.cell(row,2).value = "recordModel"
                sheet.cell(row,3).value = "insert"
                sheet.cell(row,4).value = "post"
                sheet.cell(row,5).value = "result/insert"
                sheet.cell(row,6).value = str(parm)
                row = row+1
        wb.save(self.file_name)
    def readParmfromExcel(self):
        wb = load_workbook(self.file_name)
        ws = wb.active
        sheet = wb[self.sheet]
        parmdata = []
        for i in range(2, sheet.max_row + 1):

            parmdata.append(sheet.cell(i, 6).value)
        return parmdata

    def writeResponsetoExcel(self, row,response):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        parmdata = DoExcel(project_path.do_excel_new3_path, "Parm").readParmfromExcel()
        for i in parmdata:
            sheet.cell(row, 7).value = response
            row = row+1
            wb.save(self.file_name)


# if __name__ == '__main__':
# #     #实例化
#     path=project_path.do_excel_new3_path
#     testers=DoExcel(path, "Sheet1")
#     items_scores=DoExcel(path, "Sheet2")
#     types=DoExcel(path, "Sheet3")
#     print(items_scores.read_itemsAndScore()[0])
#     print(items_scores.read_itemsAndScore()[1])
#     print(testers.read_testers())
#     print(types.read_types())
#     parm=DoExcel(path, "Parm")
#     parm.writeParmtoExcel(2)
#     row = len(items_scores.read_itemsAndScore()[0])
#     print(row)
#     DoExcel(path, "Parm").writeResponsetoExcel(2,"repose")
#     data = DoExcel(path, "Parm").readParmfromExcel()
#     print(data[0])











