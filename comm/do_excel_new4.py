from openpyxl import load_workbook
from conf import project_path
import time
class DoExcel():
    def __init__(self,file_name,sheet):
        self.file_name=file_name
        self.sheet=sheet
    def no_reg_tel(self):
        wb = load_workbook(self.file_name)
        sheet = wb['init']
        no_reg_tel=sheet.cell(1,2).value
        return no_reg_tel
    def read_excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        id_data=[]
        for i in range(2, sheet.max_row + 1):
            id_data.append(sheet.cell(i, 1).value)
        return id_data
    def read_excel1(self):
        wb=load_workbook(self.file_name)
        ws = wb.active
        sheet=wb[self.sheet]
        test_type = "1"
        scores = "120"
        item_id = "17"
        tester_id = "1211"
        eva_Type = "1"
        keyNum = "1"
        id_data=[]
        for i in range(2, sheet.max_row + 1):
            list_data =[]
            for j in  range(1,7):
                if(i==3 and j==6):
                    parm_date =eval(sheet.cell(i, 6).value)
                    #print(type(parm_date))
                    #print((parm_date))
                    if parm_date["testType"] == "test_type":
                        parm_date["testType"] = test_type
                    if parm_date["score"] == "scores":
                        parm_date["score"] = scores
                    if parm_date["itemId"] == "item_id":
                        parm_date["itemId"] = item_id
                    if parm_date["testerId"] == "tester_id":
                        parm_date["testerId"] = tester_id
                    if parm_date["keyNum"] == "keyNum":
                        parm_date["keyNum"] = keyNum
                    if parm_date["evaType"] == "eva_Type":
                        parm_date["evaType"] = eva_Type
                    if parm_date["uuid"] == "uuid":
                        parm_date["uuid"] = time.time()
                    if parm_date["startDate"] == "startDate":
                        parm_date["startDate"] = time.strftime("%Y-%m-%d %H:%M:%S")
                    if parm_date["endDate"] == "endDate":
                        parm_date["endDate"] = time.strftime("%Y-%m-%d %H:%M:%S")
                    print(parm_date)
                    list_data.append(parm_date)
                list_data.append(sheet.cell(i, j).value)
            id_data.append(list_data)
        return id_data
    def write_excel(self,row,actual,result):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet]
        sheet.cell(row, 8).value = actual
        sheet.cell(row, 9).value = result
        wb.save(self.file_name)

if __name__ == '__main__':
    #实例化
    path="D:\\study\pyProj\\test_data\\test4.xlsx"
    #print(path)
    t=DoExcel(path, "Sheet4")
    print(t.read_excel1())
    #print(type(t.read_excel1()))











