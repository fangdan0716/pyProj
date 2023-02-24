import json

import unicodedata
from openpyxl import load_workbook
wb=load_workbook("D:\\study\\pyProj\\test_data\\工作簿1.xlsx")
class DoExcel():
    def __init__(self,file_name,sheet):
        self.file_name=file_name
        self.sheet=sheet
    def read_excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet]
        test_data=[]

        for i in range(2,sheet.max_row+1):
            list_data=[]
            for j in range(1,6):
                list_data.append(sheet.cell(i, j).value)
            test_data.append(list_data)
        return test_data

    def excel_change_json(self,test_data):
        data =[]
        test_data = self.read_excel()
        print(len(test_data))
        for i in range(0,len(test_data)):
            dict1 = {}
            for j in range(0,len(test_data[i])):
                dict1["testerName"] = test_data[i][0]
                dict1["testerNumber"] = test_data[i][1]
                dict1["gender"] = test_data[i][2]
                dict1["grade"] = test_data[i][3]
                dict1["clazz"] = test_data[i][4]
            data.append(dict1)
        return data


if __name__ == '__main__':
    t = DoExcel("D:\\study\\pyProj\\test_data\\工作簿1.xlsx", "Sheet1")
    test_data = t.read_excel()
    print(test_data)

    print(t.excel_change_json(test_data))
    #json.dumps(dict1)
    #test_data = t.excel_change_json(t.read_excel())
    #print(test_data)
    data =t.excel_change_json(test_data)
    list=[]
    for i in range(len(data)):
        print(json.dumps(data[i]))
        list.append(json.dumps(data[i]))
    print(list)
    # for j in  range(len(list)):
    #     #print(type(list[j]))
    #     dict1 = json.loads(list[j])
    #     print (json.loads('"%s"' %dict1["testerName"]))